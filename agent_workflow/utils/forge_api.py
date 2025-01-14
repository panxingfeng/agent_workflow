import os

import pandas as pd

from config.tool_config import FORGE_SDWEBUI_USERNAME, FORGE_SDWEBUI_PASSWORD, FORGE_SDWEBUI_PORT

from typing import List, Tuple, Dict, Any
import requests


class ForgeAPI:
    def __init__(self, host: str = f"http://127.0.0.1:{FORGE_SDWEBUI_PORT}", username: str = FORGE_SDWEBUI_USERNAME, password: str = FORGE_SDWEBUI_PASSWORD):
        """
        初始化ForgeAPI客户端

        Args:
            host: API服务器地址
            username: 用户名（可选）
            password: 密码（可选）
        """
        self.host = host.rstrip("/")
        self.auth = (username, password) if username and password else None
        self.session = requests.Session()
        if self.auth:
            self.session.auth = self.auth

    def _get(self, endpoint: str) -> dict:
        """
        发送GET请求到API

        Args:
            endpoint: API端点

        Returns:
            响应数据
        """
        response = self.session.get(f"{self.host}{endpoint}")
        if response.status_code == 404:
            return {}
        response.raise_for_status()
        return response.json()

    def _post(self, endpoint: str, json_data: dict) -> dict:
        """
        发送POST请求到API

        Args:
            endpoint: API端点
            json_data: 请求数据

        Returns:
            响应数据
        """
        response = self.session.post(f"{self.host}{endpoint}", json=json_data)
        response.raise_for_status()
        return response.json()

    def get_models(self) -> List[str]:
        """
        获取所有可用的SD模型名称

        Returns:
            模型名称列表
        """
        try:
            models = self._get("/sdapi/v1/sd-models")
            if not models:
                print("API返回为空")
                return []

            model_names = [model.get('model_name', '') for model in models]

            return model_names

        except Exception as e:
            print(f"获取模型列表时发生错误: {str(e)}")
            return []

    def get_loras(self) -> List[str]:
       """
       获取所有可用的Lora模型名称

       Returns:
           模型名称列表
       """
       loras = self._get("/sdapi/v1/loras")
       return [lora.get('name', '') for lora in loras]

    def add_loras_to_prompt(self, prompt: str, lora_names: List[str],
                            excel_path: str = os.path.join(os.getcwd(), "config", "lora_model_info.xlsx")) -> tuple[
                                                                                                                  str, None] | str:
        """
        添加多个Lora模型到提示词，从Excel文件读取对应的权重和触发词

        Args:
            prompt: 原始提示词
            lora_names: Lora模型名称列表
            excel_path: 配置Excel文件的路径
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            lora_config = {}
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2):
                values = [cell.value for cell in row]
                if values[0]:
                    row_data = dict(zip(headers, values))
                    lora_config[str(values[0])] = row_data

        except Exception as e:
            print(excel_path)
            print(f"读取lora_model_info的Excel文件失败: {e}")
            # 如果读取失败，为每个lora生成一个标签
            lora_parts = [f"<lora:{name}:1>" for name in lora_names]
            return f"{' '.join(lora_parts)}, {prompt}", None

        lora_parts = []
        trigger_words = []
        base_models = []  # 收集所有的base_model

        # 为每个lora名称创建单独的标签
        for name in lora_names:
            if name in lora_config:
                row = lora_config[name]

                try:
                    weight = float(row['weight']) if row['weight'] not in [None, '', '<unset>'] else 1.0
                except:
                    weight = 1.0

                trigger_word = str(row.get('trigger_word', '')).strip()
                if trigger_word in ['None', '<unset>', 'nan']:
                    trigger_word = ''

                lora_prompt_name = str(row.get('lora_prompt_name', '')).strip()
                if lora_prompt_name in ['None', '<unset>', 'nan']:
                    lora_prompt_name = ''

                actual_name = lora_prompt_name if lora_prompt_name else name

                # 添加lora标签
                lora_parts.append(f"<lora:{actual_name}:{weight}>")

                # 如果有触发词则添加到列表
                if trigger_word:
                    trigger_words.append(trigger_word)
            else:
                # 如果找不到配置，使用默认权重1.0
                lora_parts.append(f"<lora:{name}:1>")

        # 组合最终的prompt
        result = f"{' '.join(lora_parts)}"
        if trigger_words:
            result += f" {', '.join(trigger_words)}"
        if prompt:
            result += f", {prompt}"

        return result

    def get_model_config(self, model_name: str,
                         excel_path: str = os.path.join(os.getcwd(), "config", "base_model_info.xlsx")) -> dict[
                                                                                                               str, Any] | None:
        """
        根据模型名称获取配置参数

        Args:
            model_name: 模型名称
            excel_path: 配置Excel文件路径

        Returns:
            dict: 包含模型配置的字典
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            # 查找匹配的行
            for row in ws.iter_rows(min_row=2):
                values = [cell.value for cell in row]
                if values[0] and values[0] == model_name:
                    config = {
                        'model_name': values[0],
                        'sampling_method': values[1],
                        'steps': values[2],
                        'width': values[3],
                        'height': values[4],
                        'cfg_scale': values[5],
                        'trigger_word': values[6]
                    }
                    return config

            print(f"未找到模型 {model_name} 的配置信息")
            return None

        except Exception as e:
            print(f"读取配置文件失败: {e}")
            return None


# print("支持的基础模型:" + str(ForgeAPI().get_models()))
# print("支持的lora模型:" + str(ForgeAPI().get_loras()))